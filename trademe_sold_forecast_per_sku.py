import openpyxl
from datetime import datetime, timedelta

def process_sales_data(input_file, output_file):
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        # 字典用于存储每个SKU的第一个售出日期和最后一个售出日期以及总销售数量
        sku_data = {}

        # 遍历表格中的数据
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku, date_str, quantity = row[39], row[3], row[8]

            if sku not in sku_data:
                sku_data[sku] = {'first_sale_date': datetime.max, 'last_sale_date': datetime.min, 'total_quantity': 0}

            # 将日期字符串转换为datetime对象
            date_obj = datetime.strptime(date_str, '%b %d %Y %I:%M%p')
            sku_data[sku]['total_quantity'] += quantity

            # 更新最早和最晚的销售日期
            if date_obj < sku_data[sku]['first_sale_date']:
                sku_data[sku]['first_sale_date'] = date_obj
            if date_obj > sku_data[sku]['last_sale_date']:
                sku_data[sku]['last_sale_date'] = date_obj

        # 创建一个新的工作簿和工作表
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active

        # 写入表头
        result_ws.append(['SKU', '第一个售出日期', '最后一个售出日期', '总销售数量', '每天平均销售数量', '每月平均销售数量'])

        # 遍历每个SKU的数据，计算并写入结果
        for sku, data in sku_data.items():
            first_sale_date = data['first_sale_date'].strftime('%Y-%m-%d')
            last_sale_date = data['last_sale_date'].strftime('%Y-%m-%d')
            total_quantity = data['total_quantity']
            days_diff = (data['last_sale_date'] - data['first_sale_date']).days + 1
            daily_average = total_quantity / days_diff

            # 添加条件判断，根据总销售数量确定每月平均销售数量
            if total_quantity <= 2:
                monthly_average = total_quantity
            else:
                monthly_average = daily_average * 30

            result_ws.append([sku, first_sale_date, last_sale_date, total_quantity, daily_average, monthly_average])

        # 保存结果到新的Excel文件
        result_wb.save(output_file)
        print(f'结果已保存到 "{output_file}"')

    except FileNotFoundError:
        print(f'找不到文件 "{input_file}"，请确保路径和文件名正确并存在。')

    except Exception as e:
        print(f'发生错误: {e}')

# 调用函数，传入输入文件和输出文件的路径
input_file_path = 'C:\\Users\\ThinkPad\\SynologyDrive\\Trademe\\sold.xlsx'
output_file_path = 'C:\\Users\\ThinkPad\\SynologyDrive\\Trademe\\result.xlsx'
process_sales_data(input_file_path, output_file_path)

# 导出trademe销售数据表格, 命名为sold, 放到synology根目录下
# 自动统计每个月的销量, 并将结果储存到result.xlsx, 供后续程序做库存预警使用
# 用法: call function process_sales_data(input_file_path, output_file_path)